"""
Run this script whenever new data is available.
Splits it from a .xlsx Workbook to .csv files

"""
from openpyxl import load_workbook
from pathlib import Path
from csv import writer

FILENAMES = {
	"rg_course": "courses",
	"rg_section": "section",
	"rg_sched": "schedule",
	"rg_sect_sched": "section_schedule",
	"students": "students",
	"faculty": "faculty",
}

data_dir = Path.cwd().parent / "data"

def get_worksheet(workbook, sheet_name): 
	for sheet in workbook.sheetnames: 
		if sheet.lower() == sheet_name: 
			return workbook [sheet]
	else: 
		raise KeyError(f"Could not find sheet {sheet_name} in the data")

def convert_to_csv(sheet, filename): 
	with open(filename, "w", newline = "") as file: 
		csv = writer(file)
		for row in sheet.rows: 
			csv.writerow([cell.value for cell in row])

if __name__ == "__main__": 
	workbook = load_workbook(data_dir / "data.xlsx")
	for sheet_name, filename in FILENAMES.items():
		convert_to_csv(
			get_worksheet(workbook, sheet_name), 
			data_dir / f"{filename}.csv"
		)